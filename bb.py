import json, re 
import openpyxl # xlsx 库
from openpyxl.styles import PatternFill
# 文件列表和文件路径
file_list = ['zh-tw', 'fr-ca',  'in-id']
auth_list = ["auth", "myalert"]
workbook = openpyxl.load_workbook(f'D:/project/python/test.xlsx') # 读取Excel文件
worksheet = workbook.active # excel表
# 遍历1：AUTH文件列表
for auth in auth_list:
    en_gb_json_filepath = f'D:/project/python/{auth}/en-gb.json'
    if 1:  # 启用容错(如果文件不存在、跳过)
        try:
            with open(en_gb_json_filepath, 'r', encoding='utf-8') as f:
                en_gb_json = json.load(f)
        except FileNotFoundError:
            print(f"文件 {en_gb_json_filepath} 不存在。")
            continue
    with open(en_gb_json_filepath, 'r', encoding='utf-8') as f:
        en_gb_json = json.load(f)
        en_gb_json_processed = {k: v.replace(
            '\\n', '\n') for k, v in en_gb_json.items()}
    # 遍历2：各国json文件列表
    for file in file_list:
        json_file_path = f'D:/project/python/{file}.json'
        if 1:  # 启用容错
            try:
                with open(json_file_path, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)
            except FileNotFoundError:
                print(f"文件 {json_file_path} 不存在。")
                continue
        with open(json_file_path, 'r', encoding='utf-8') as f2:
            json_data = json.load(f2)
            json_data_processed = {k2: v2.replace(
                '\\n', '\n') for k2, v2 in json_data.items()}

        # 先根据excel表第一行标题来创建一个字典存入每个语言和其对应的列索引
        index_dict = {cell.value: i for i, cell in enumerate(worksheet[1])} # 类似 { "en-gb": 0, "zh-tw": 1, ... }
        # 遍历3：Excel当前表的每一行
        for row in worksheet.iter_rows(min_row=2):
            en_key = row[0].value  # "Review 1 selected \n request"
            en_key_processed = en_key.replace('\\n', '\n') # "Review 1 selected \n request"，尾部带有 _processed 字样的属于去掉\n的，下同
            value = row[index_dict[file]].value # "Examiner une \n demande sélectionnée一致" 即法语
            value__processed = value.replace('\\n', '\n') # "Examiner une demande sélectionnée一致"  即法语去掉\n
            # 如果en_key在en_gb_json的值中 （即字典 ['Review 1 selected \n request','Review 2 selected \n request']中）即找到'Review 2 selected \n request'
            if en_key_processed in list(en_gb_json_processed.values()):
                nls_key_to_find = [
                    k for k, v in en_gb_json_processed.items() if v == en_key_processed][0]  # 即 NLS KEY = "AUTH.RFP.REVIEW_ONE_REQUEST"
                ### json_data[nls_key_to_find] = value.replace('\\n', '\n') #直接更新当前国json（不能直接更新、以下改为判断英/法）
  
                # 1111111111111 填充excel的H列
                for cell in row:
                    cell_address = cell.coordinate  # 获取当前单元格的地址（例如"A1"）
                    print('cell_address', cell_address)
                    if cell_address == '!ref':
                        continue  # 跳过空格
                    if cell.value is not None:  
                        cell_value_processed = cell.value.replace('\\n', '\n')  
                    else:  
                        cell_value_processed = cell.value
                    key_value_processed = en_gb_json.get(nls_key_to_find).replace('\\n', '\n')
                    # 查找A列中的英语匹配项（必须去掉\\n来进行对比）
                    if cell_value_processed == key_value_processed and re.search(r'[A-Z]+', cell_address).group() == 'A':
                        print('find a cell at col "A" that is en ----', cell.value)
                        row_number = int(re.search(r'\d+', cell_address).group())
                        col_letter = re.search(r'[A-Z]+', cell_address).group()
                        worksheet['H' + str(row_number)].value = nls_key_to_find
  
                # 222222222222222 如果英语json和法语json值相等
                if en_gb_json_processed[nls_key_to_find] == json_data_processed[nls_key_to_find]:
                    print('find that two json is the same ======', en_gb_json_processed[nls_key_to_find]) 
                    print('find that two json is the same ======', value__processed) 
                    json_data_processed[nls_key_to_find]=value__processed
                    json_data[nls_key_to_find]=value
                # 3333333333333 如果英语json和法语json值不相等、且法语也不相等
                elif en_gb_json_processed[nls_key_to_find] != json_data_processed[nls_key_to_find] and json_data_processed[nls_key_to_find] != value__processed:
                    fill = PatternFill(fill_type="solid", fgColor="FFFF00") # 创建一个填充样式
                    for cell2 in row:
                        # print(cell2)
                        cell2_address = cell2.coordinate  # 获取当前单元格的地址（例如"A1"）
                        # print('cell2_address', cell2_address)
                        # 如果单元格的值等于 frCaValue，设置单元格的填充样式
                        if cell2.value == value:
                            cell2.fill = fill
                else:
                    print('the same')

        # 将更新后的json_data写入文件
        with open(json_file_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)

workbook.save(f'D:/project/python/new_after.xlsx')
